# -*- coding: utf-8 -*-
"""
Created on Sun Nov 19 12:21:16 2017

"""

# Handles the HTML files, providing parsing functionality
from bs4 import BeautifulSoup
# Handles operating system functions, like opening/closing a file, and reading file names from the directory
import os
# Handles the Excel files and provides parsing functionality
import openpyxl


# Handles the main application function calls
def MainInterface():
    
    acronyms = ReadExcel()
    
    print("Welcome to Lauren's Acronym Pass")
    
    filePath = "SCO_1/storyboards/"
    
    while True:
        print("")
        print("Enter a command from the list below:")
        print("1 | Search For Acronyms and Spellouts")
        print("2 | Search For Acronyms")
        print("3 | Search For Spellouts")
        print("0 | Quit")
        print("What do you want to do?")
        action = input("Enter input here... ")
    
        if action == "1":
           for filename in os.listdir(filePath): 
               print("")
               print(filename + ": ")
               SearchForTagAcronyms('p', FileHandler(filename), acronyms)
               SearchForTagAcronyms('li', FileHandler(filename), acronyms)
               SearchForTagAcronyms('title', FileHandler(filename), acronyms)
               ScanForSpellouts('p', FileHandler(filename), acronyms)
               ScanForSpellouts('li', FileHandler(filename), acronyms)
               ScanForSpellouts('title', FileHandler(filename), acronyms)
               print("")
           WaitToContinue()
        elif action == "2":
            for filename in os.listdir(filePath):
                print("")
                print(filename + ": ")
                SearchForTagAcronyms('p', FileHandler(filename), acronyms)
                SearchForTagAcronyms('li', FileHandler(filename), acronyms)
                SearchForTagAcronyms('title', FileHandler(filename), acronyms)
                print("")
            WaitToContinue()
        elif action == "3":
           for filename in os.listdir(filePath): 
               print("")
               print(filename + ": ")
               ScanForSpellouts('p', FileHandler(filename), acronyms)
               ScanForSpellouts('li', FileHandler(filename), acronyms)
               ScanForSpellouts('title', FileHandler(filename), acronyms)
               print("")
           WaitToContinue()
           
        elif action == "0":
            Quit()
            break
        else:
            print("Invalid input. Try again.")
            print("")

# Reads in an Excel .xlsx spreadsheet, parsing through the rows, and saving each row in a dictionary
def ReadExcel():
    
    wb = openpyxl.load_workbook('Acros Final Project.xlsx')
    sheet = wb.get_sheet_by_name('Acronyms')
    sheet = wb.active
    
    acroSpellDict = {}
    
    #read in data from Excel sheet
    for i in range(1, sheet.max_row):
        # store i row's acronym in tempAcro
        tempAcro = sheet.cell(row=i, column=1).value
        # store i row's spellout in tempSpell
        tempSpell = sheet.cell(row=i, column=2).value
        # add the acroynm and spellout to the dictionary
        acroSpellDict.update({tempAcro: tempSpell})
        
    # return the dictionary storing the acronyms and spellout
    return acroSpellDict

# Using the acros dictionary, the function searches for each spellout by using the SearchForTagText function
def ScanForSpellouts(t, w, acros):
    
    for txt in acros.values():
        SearchForTagText(t, w, txt)
    
    print("")
    
# Searches for a specified string within the webpage, and prints out what it finds
def SearchForTagText(tag, webpage, text):
    
    paragraphs = webpage.select(tag)

    for para in paragraphs:
        if str(text) in str(para):
            print("The spellout '" + text + "' was found.")
    
    return webpage

# Searches for the specified acronym from the acros dictionary, printing out what it finds
def SearchForTagAcronyms(tag, webpage, acros):
    
    paragraphs = webpage.select(tag)

    for para in paragraphs:
        tempString = para.getText()
        tempString = str(para)
        for key in acros.keys():
            if str(key) in str(tempString):
                print("The acronym " + str(key) + " was found.")    
    
    return webpage

# Prompts the user to enter any key to continue with the application.
def WaitToContinue():
    
    input("Enter any key to continue...")
    print("")
 
# Quits the application    
def Quit():
    
    print("Quitting application...Goodbye.")

# Opens the specified file, creates a BeautifulSoup page, and closes the file.
def FileHandler(fileName):

    filePath = "SCO_1/storyboards/" + fileName
    
    file = open(filePath, "r")
    
    page = BeautifulSoup(file, "lxml")
    
    file.close()
    
    return page    
    
# Starts the application
MainInterface()

        
